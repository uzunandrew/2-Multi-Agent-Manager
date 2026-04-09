import sys, os, io, json, argparse
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# === Аргументы ===
parser = argparse.ArgumentParser(description='Генерация Excel-отчёта аудита')
parser.add_argument('project_path', help='Путь к папке проекта (содержит project_info.json и _output/)')
args = parser.parse_args()

project_path = args.project_path.rstrip('/\\')
output_dir = os.path.join(project_path, '_output')

# === Чтение данных ===
with open(os.path.join(project_path, 'project_info.json'), 'r', encoding='utf-8') as f:
    project_info = json.load(f)

project_id = project_info['project_id']
section = project_info.get('section', '')

with open(os.path.join(output_dir, 'findings.json'), 'r', encoding='utf-8') as f:
    fdata = json.load(f)
with open(os.path.join(output_dir, 'optimization.json'), 'r', encoding='utf-8') as f:
    odata = json.load(f)
with open(os.path.join(output_dir, 'optimization_review.json'), 'r', encoding='utf-8') as f:
    ordata = json.load(f)

findings = fdata['findings']
proposals = odata.get('proposals', [])
opt_verdicts = {r['opt_id']: r for r in ordata.get('reviews', [])}

# === Стили ===
wb = Workbook()
now = datetime.now().strftime('%d.%m.%Y %H:%M')
thin = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))

HEADER_FILL = PatternFill('solid', fgColor='1F497D')
HEADER_FONT = Font(bold=True, color='FFFFFF', size=10)
TITLE_FILL = PatternFill('solid', fgColor='2E75B6')
TITLE_FONT = Font(bold=True, color='FFFFFF', size=11)
TOTAL_FILL = PatternFill('solid', fgColor='2E3F50')

CAT_MAP = {
    'Критическое':      {'emoji': '\U0001f534', 'label': 'КРИТИЧЕСКОЕ',      'row_fill': 'FFCCCC', 'cell_fill': 'FFD9D9', 'font_color': 'C00000'},
    'Экономическое':    {'emoji': '\U0001f7e0', 'label': 'ЭКОНОМИЧЕСКОЕ',    'row_fill': 'FCE4D6', 'cell_fill': 'FDEBD9', 'font_color': 'C05000'},
    'Эксплуатационное': {'emoji': '\U0001f7e1', 'label': 'ЭКСПЛУАТАЦИОННОЕ', 'row_fill': 'FFFACD', 'cell_fill': 'FFFFB3', 'font_color': '806600'},
}

def ci(cat):
    return CAT_MAP.get(cat, {'emoji': '\u26aa', 'label': cat.upper(), 'row_fill': 'F8F8F8', 'cell_fill': 'F8F8F8', 'font_color': '333333'})

cat_order = ['Критическое', 'Экономическое', 'Эксплуатационное']

counts = {}
for f in findings:
    counts[f['category']] = counts.get(f['category'], 0) + 1

# ===== СВОДКА =====
ws1 = wb.active
ws1.title = 'СВОДКА'
h1 = ['№', 'Проект (ID)', 'Объект / Раздел']
for cat in cat_order:
    h1.append(f'{ci(cat)["emoji"]} {cat[:6]}.')
h1.append('Итого')
w1 = [5, 30, 38, 14, 14, 16, 10]
for i, (h, w) in enumerate(zip(h1, w1), 1):
    c = ws1.cell(row=1, column=i, value=h)
    c.font = HEADER_FONT; c.fill = HEADER_FILL; c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True); c.border = thin
    ws1.column_dimensions[get_column_letter(i)].width = w

ncols = len(h1)
ws1.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
c = ws1.cell(row=2, column=1, value=f'СВОДНЫЙ ОТЧЁТ АУДИТА   |   Дата: {now}   |   Мультиагентный пайплайн')
c.font = TITLE_FONT; c.fill = TITLE_FILL
for col in range(1, ncols + 1): ws1.cell(row=2, column=col).border = thin

row3 = [1, project_id, f'{project_id} / {section}'] + [counts.get(cat, 0) for cat in cat_order] + [len(findings)]
for i, v in enumerate(row3, 1):
    c = ws1.cell(row=3, column=i, value=v)
    c.font = Font(bold=(i in [2, ncols]), size=10); c.alignment = Alignment(horizontal='center', vertical='center'); c.border = thin
    if 4 <= i <= 3 + len(cat_order): c.fill = PatternFill('solid', fgColor=ci(cat_order[i-4])['cell_fill'])

for col in range(1, ncols + 1):
    c = ws1.cell(row=4, column=col)
    c.fill = TOTAL_FILL; c.font = Font(bold=True, color='FFFFFF', size=10); c.alignment = Alignment(horizontal='center'); c.border = thin
ws1.cell(row=4, column=1, value='ИТОГО')
for i, cat in enumerate(cat_order, 4):
    ws1.cell(row=4, column=i, value=counts.get(cat, 0))
ws1.cell(row=4, column=ncols, value=len(findings))

# ===== ЗАМЕЧАНИЯ =====
ws2 = wb.create_sheet(project_id)
h2 = ['№', 'Лист/Раздел', 'Проблема', 'Описание', 'Решение', 'Категория', 'Чем грозит']
w2 = [5, 24, 28, 52, 48, 22, 32]
for i, (h, w) in enumerate(zip(h2, w2), 1):
    c = ws2.cell(row=1, column=i, value=h)
    c.font = HEADER_FONT; c.fill = HEADER_FILL; c.alignment = Alignment(horizontal='center' if i == 1 else 'left', vertical='center', wrap_text=True); c.border = thin
    ws2.column_dimensions[get_column_letter(i)].width = w

ws2.merge_cells('A2:G2')
c = ws2.cell(row=2, column=1, value=f'{project_id}  |  аудит от {now}  |  замечаний: {len(findings)}')
c.font = TITLE_FONT; c.fill = TITLE_FILL
for col in range(1, 8): ws2.cell(row=2, column=col).border = thin

cat_pri = {cat: i for i, cat in enumerate(cat_order)}
sf = sorted(findings, key=lambda f: cat_pri.get(f['category'], 9))

RISK = {
    'Критическое': 'Нарушение обязательных норм',
    'Экономическое': 'Финансовые потери при закупке/монтаже',
    'Эксплуатационное': 'Проблемы при эксплуатации объекта',
}

for idx, f in enumerate(sf):
    r = idx + 3
    info = ci(f['category'])
    rf = PatternFill('solid', fgColor=info['row_fill'])
    sheet_ref = f'Лист {f["sheet"]}' if f.get('sheet') else f'Стр. {f.get("page", "")}'
    vals = [idx+1, sheet_ref, f.get('title',''), f.get('description',''), f.get('recommendation',''), f'{info["emoji"]} {info["label"]}', RISK.get(f['category'], '')]
    for col, v in enumerate(vals, 1):
        c = ws2.cell(row=r, column=col, value=v)
        c.fill = rf; c.border = thin
        c.alignment = Alignment(horizontal='center' if col == 1 else 'left', vertical='top', wrap_text=True)
        if col == 6: c.font = Font(bold=True, color=info['font_color'], size=10)
        elif col == 4: c.font = Font(size=9)
        elif col == 1: c.font = Font(bold=True, size=10)
        else: c.font = Font(size=10)

rf = len(sf) + 3
ws2.merge_cells(f'A{rf}:G{rf}')
ws2.cell(row=rf, column=1, value=f'Итого замечаний: {len(findings)}  |  Мультиагентный аудит  |  {now}').font = Font(italic=True, size=9, color='666666')
ws2.cell(row=rf, column=1).alignment = Alignment(horizontal='center')

# ===== ОПТИМИЗАЦИЯ =====
ws3 = wb.create_sheet(f'ОПТ {project_id}')
h3 = ['№', 'ID', 'Категория', 'Текущее решение', 'Предложение', 'Экономия', 'Вердикт', 'Привязка', 'Риски']
w3 = [5, 10, 24, 42, 42, 24, 16, 12, 32]
for i, (h, w) in enumerate(zip(h3, w3), 1):
    c = ws3.cell(row=1, column=i, value=h)
    c.font = HEADER_FONT; c.fill = HEADER_FILL; c.alignment = Alignment(horizontal='center' if i <= 2 else 'left', vertical='center', wrap_text=True); c.border = thin
    ws3.column_dimensions[get_column_letter(i)].width = w

passed = sum(1 for r in ordata.get('reviews', []) if 'pass' in r.get('verdict', ''))
ws3.merge_cells('A2:I2')
c = ws3.cell(row=2, column=1, value=f'{project_id}  |  оптимизаций: {len(proposals)} (принято: {passed})  |  отчёт: {now}')
c.font = TITLE_FONT; c.fill = TITLE_FILL
for col in range(1, 10): ws3.cell(row=2, column=col).border = thin

TYPE_E = {
    'замена_материалов':             '\U0001f4b0 Замена материалов',
    'оптимизация_монтажа':           '\U0001f527 Оптимизация монтажа',
    'упрощение_конструкций':         '\U0001f3d7\ufe0f Упрощение конструкций',
    'схемные_решения':               '\u26a1 Схемные решения',
    'оптимизация_трасс':             '\U0001f4cf Оптимизация трасс',
    'технологические_альтернативы':  '\U0001f680 Технологические альтернативы',
    'нагрузки_и_запасы':             '\U0001f4ca Нагрузки и запасы',
}
VF = {
    'pass': PatternFill('solid', fgColor='D4EDDA'),
    'pass_with_conditions': PatternFill('solid', fgColor='FFF3CD'),
    'reject_norm_violation': PatternFill('solid', fgColor='F8D7DA'),
    'reject_reliability': PatternFill('solid', fgColor='F8D7DA'),
    'conflict_with_finding': PatternFill('solid', fgColor='F8D7DA'),
    'reject_not_in_vendor_list': PatternFill('solid', fgColor='F8D7DA'),
}

for idx, p in enumerate(proposals):
    r = idx + 3
    oid = p.get('opt_id', '')
    rv = opt_verdicts.get(oid, {})
    verdict = rv.get('verdict', '')
    cat_label = TYPE_E.get(p.get('category', ''), p.get('category', ''))
    vals = [idx+1, oid, cat_label, p.get('current',''), p.get('proposed',''),
            p.get('savings_estimate', ''), verdict,
            p.get('finding_ref','') or '\u2014', p.get('risk','') or '\u2014']
    for col, v in enumerate(vals, 1):
        c = ws3.cell(row=r, column=col, value=v)
        c.font = Font(size=10, bold=(col <= 2)); c.alignment = Alignment(horizontal='center' if col <= 2 else 'left', vertical='top', wrap_text=True); c.border = thin
        if col == 7 and verdict in VF: c.fill = VF[verdict]

rf = len(proposals) + 3
ws3.merge_cells(f'A{rf}:I{rf}')
ws3.cell(row=rf, column=1, value=f'Итого: {len(proposals)} предложений  |  {now}').font = Font(italic=True, size=9, color='666666')
ws3.cell(row=rf, column=1).alignment = Alignment(horizontal='center')

# === Сохранение ===
out_name = f'audit_report_{datetime.now().strftime("%d.%m.%Y")}.xlsx'
out = os.path.join(output_dir, out_name)
wb.save(out)
print(f'OK: {out}')
print(f'Findings: {len(findings)}, Optimizations: {len(proposals)}')
