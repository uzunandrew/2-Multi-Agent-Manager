#!/usr/bin/env python3
"""
Сравнение прогонов аудита проекта 133_23-ГК-ГРЩ.

Читает данные 3 прогонов из _output/ и _output_old/,
генерирует Excel с 4 листами: Сводка, Качество, Детали замечаний, Хронология.

Использование:
    python gen_comparison.py
"""

import json
import os
import sys
from datetime import datetime
from pathlib import Path
from glob import glob

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl не установлен. Установите: pip install openpyxl")
    sys.exit(1)


# ─── Пути ────────────────────────────────────────────────────────────────────

SCRIPT_DIR = Path(__file__).parent
PROJECT_DIR = SCRIPT_DIR.parent / "projects" / "EOM" / "133_23-ГК-ГРЩ"
OUTPUT = PROJECT_DIR / "_output"
OUTPUT_OLD = PROJECT_DIR / "_output_old"
EXCEL_PATH = SCRIPT_DIR / "audit_runs_comparison.xlsx"


# ─── Стили ───────────────────────────────────────────────────────────────────

HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
LABEL_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
LABEL_FONT = Font(bold=True, size=10)
RUN_HEADER_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
RUN_HEADER_FONT = Font(bold=True, size=11, color="375623")
SECTION_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
SECTION_FONT = Font(bold=True, size=10)

CAT_FILLS = {
    "Критическое": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    "КРИТИЧЕСКОЕ": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    "Экономическое": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    "ЭКОНОМИЧЕСКОЕ": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    "Эксплуатационное": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "РЕКОМЕНДАТЕЛЬНОЕ": PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid"),
}

THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def style_cell(cell, fill=None, font=None, align_h="left", wrap=False):
    cell.border = THIN_BORDER
    cell.alignment = Alignment(horizontal=align_h, vertical="center", wrap_text=wrap)
    if fill:
        cell.fill = fill
    if font:
        cell.font = font


# ─── Загрузка JSON ───────────────────────────────────────────────────────────

def load_json(path):
    if not path.exists():
        return None
    try:
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    except (json.JSONDecodeError, ValueError):
        return None


def count_findings_in_partials(directory, prefix="partial_"):
    """Считает замечания во всех partial_*.json файлах."""
    total = 0
    agents = {}
    for p in sorted(directory.glob(f"{prefix}*.json")):
        data = load_json(p)
        if data is None:
            continue
        name = p.stem.replace(prefix, "")
        if isinstance(data, dict):
            findings = data.get("findings", [])
        elif isinstance(data, list):
            findings = data
        else:
            findings = []
        agents[name] = len(findings)
        total += len(findings)
    return total, agents


def count_by_category(findings, cat_key="category", sev_key="severity"):
    """Подсчитывает замечания по категориям."""
    cats = {}
    for f in findings:
        cat = f.get(cat_key) or f.get(sev_key) or "?"
        cats[cat] = cats.get(cat, 0) + 1
    return cats


def avg_confidence(findings):
    vals = [f["confidence"] for f in findings if "confidence" in f and f["confidence"] is not None]
    if not vals:
        # practicality_score может быть в quality.practicality_score (Run 1)
        for f in findings:
            ps = None
            if "practicality_score" in f and f["practicality_score"] is not None:
                ps = f["practicality_score"]
            elif isinstance(f.get("quality"), dict):
                ps = f["quality"].get("practicality_score")
            if ps is not None:
                vals.append(ps / 100 if ps > 1 else ps)  # нормализуем 0-100 → 0-1
    return round(sum(vals) / len(vals), 3) if vals else 0


def count_opt_verdicts(reviews):
    """Подсчитывает вердикты оптимизации."""
    if reviews is None:
        return {}
    review_list = reviews.get("reviews", [])
    if not review_list and isinstance(reviews, list):
        review_list = reviews
    verdicts = {}
    for r in review_list:
        v = r.get("verdict", "?")
        verdicts[v] = verdicts.get(v, 0) + 1
    return verdicts


# ─── Сбор данных по прогонам ─────────────────────────────────────────────────

def collect_run1():
    """Run 1: 27.03.2026 — Python-пайплайн (данные в _output/)."""
    run = {"name": "Run 1 (27.03.2026)", "date": "27.03.2026"}
    run["architecture"] = "Python-пайплайн: text_analysis → block_analysis (Gemini) → findings_merge (Opus) → critic (GPT-5.4) → norms (Sonnet) → optimization (Opus) → opt_critic (Sonnet)"

    # Pipeline log
    plog = load_json(OUTPUT / "pipeline_log.json")
    if plog:
        stages = plog.get("stages", {})
        run["models"] = set()
        run["stages_info"] = {}
        for name, st in stages.items():
            m = st.get("model")
            if m:
                run["models"].add(m)
            started = st.get("started_at", "")
            ended = st.get("completed_at", "")
            if started and ended:
                try:
                    s = datetime.fromisoformat(started)
                    e = datetime.fromisoformat(ended)
                    run["stages_info"][name] = {
                        "duration_sec": (e - s).total_seconds(),
                        "model": m or "",
                        "input_tokens": st.get("input_tokens", 0),
                        "output_tokens": st.get("output_tokens", 0),
                    }
                except Exception:
                    pass
        run["models"] = ", ".join(sorted(run["models"]))

        # Wall-clock
        all_times = []
        for st in stages.values():
            for k in ("started_at", "completed_at"):
                if st.get(k):
                    try:
                        all_times.append(datetime.fromisoformat(st[k]))
                    except Exception:
                        pass
        if all_times:
            run["wall_clock_min"] = round((max(all_times) - min(all_times)).total_seconds() / 60, 1)

        # Tokens
        run["total_input_tokens"] = sum(st.get("input_tokens", 0) for st in stages.values())
        run["total_output_tokens"] = sum(st.get("output_tokens", 0) for st in stages.values())

    # Findings (03_findings.json — Run 1)
    f03 = load_json(OUTPUT / "03_findings.json")
    if f03:
        findings = f03.get("findings", [])
        run["findings_count"] = len(findings)
        run["findings_by_cat"] = count_by_category(findings, sev_key="severity")
        run["avg_confidence"] = avg_confidence(findings)
        run["findings_list"] = findings
        meta = f03.get("meta", {})
        run["blocks_analyzed"] = meta.get("blocks_analyzed", 0)
    else:
        run["findings_count"] = 0
        run["findings_by_cat"] = {}
        run["avg_confidence"] = 0
        run["findings_list"] = []

    # Text analysis
    ta = load_json(OUTPUT / "01_text_analysis.json")
    if ta:
        if isinstance(ta, dict):
            run["text_findings"] = len(ta.get("findings", []))
        elif isinstance(ta, list):
            run["text_findings"] = len(ta)
        else:
            run["text_findings"] = 0
    else:
        run["text_findings"] = 0

    run["r1_agents"] = 0
    run["agent_names"] = "нет (монолитный анализ)"
    run["critics_count"] = 1
    run["critics_names"] = "findings_critic (GPT-5.4)"
    run["rejected_count"] = 0
    run["raw_from_agents"] = run["text_findings"]

    # Optimization (стадия optimization в pipeline_log — данные в audit_trail)
    opt = load_json(OUTPUT / "audit_trail" / "05_optimization_2026-03-27T00-32-25.json") if (OUTPUT / "audit_trail").exists() else None
    if opt:
        opts = opt.get("optimizations", opt.get("proposals", []))
        if isinstance(opt, list):
            opts = opt
        run["opt_count"] = len(opts) if isinstance(opts, list) else 0
    else:
        run["opt_count"] = 13  # из pipeline данных

    run["opt_accepted"] = "н/д"
    run["opt_accepted_pct"] = "н/д"

    return run


def collect_run2():
    """Run 2: 28-29.03.2026 — Claude Code, 6 агентов (данные в _output_old/)."""
    run = {"name": "Run 2 (28-29.03.2026)", "date": "28-29.03.2026"}
    run["architecture"] = "Claude Code: 6 R1-агентов → 2 критика (review_a + review_b) → синтез → оптимизация"

    # Findings
    f_data = load_json(OUTPUT_OLD / "findings.json")
    if f_data:
        findings = f_data.get("findings", [])
        run["findings_count"] = len(findings)
        run["findings_by_cat"] = count_by_category(findings)
        run["avg_confidence"] = avg_confidence(findings)
        run["findings_list"] = findings
    else:
        run["findings_count"] = 0
        run["findings_by_cat"] = {}
        run["avg_confidence"] = 0
        run["findings_list"] = []

    # Raw from agents
    raw, agents = count_findings_in_partials(OUTPUT_OLD)
    run["raw_from_agents"] = raw
    run["r1_agents"] = len(agents)
    run["agent_names"] = ", ".join(sorted(agents.keys()))

    # Rejected
    rej = load_json(OUTPUT_OLD / "rejected.json")
    if rej:
        run["rejected_count"] = rej.get("total", len(rej.get("rejected", [])))
    else:
        run["rejected_count"] = 0

    # Filtered
    filt = load_json(OUTPUT_OLD / "filtered_findings.json")
    run["filtered_count"] = len(filt.get("findings", [])) if filt and isinstance(filt, dict) else (len(filt) if isinstance(filt, list) else 0)

    # Critics
    run["critics_count"] = 2
    run["critics_names"] = "review_a, review_b (Claude Code)"

    # Optimization
    opt = load_json(OUTPUT_OLD / "optimization.json")
    if opt:
        opts = opt.get("optimizations", opt.get("proposals", []))
        if isinstance(opts, list):
            run["opt_count"] = len(opts)
        else:
            run["opt_count"] = 0
    else:
        run["opt_count"] = 0

    opt_rev = load_json(OUTPUT_OLD / "optimization_review.json")
    if opt_rev:
        verdicts = count_opt_verdicts(opt_rev)
        accepted = verdicts.get("pass", 0) + verdicts.get("pass_with_conditions", 0)
        run["opt_accepted"] = accepted
        run["opt_verdicts"] = verdicts
        run["opt_accepted_pct"] = f"{round(accepted / run['opt_count'] * 100)}%" if run["opt_count"] else "—"
    else:
        run["opt_accepted"] = "н/д"
        run["opt_accepted_pct"] = "н/д"
        run["opt_verdicts"] = {}

    # Время — по timestamps файлов
    run["wall_clock_min"] = "~28ч (28.03 10:20 → 29.03 14:16)"
    run["models"] = "Claude Opus 4.6 (R1 + синтез), Claude Sonnet 4.6 (критики)"
    run["total_input_tokens"] = "н/д (Claude Code)"
    run["total_output_tokens"] = "н/д (Claude Code)"
    run["stages_info"] = {}

    return run


def collect_run3_v2():
    """Run 3 (v2): 04.04.2026 — Claude Code, сгруппированные агенты (Sonnet+Opus)."""
    run = {"name": "Run 3 v2 (04.04.2026)", "date": "04.04.2026"}
    run["architecture"] = "Claude Code: 4 группы (group_a: cables/tables/metering/power_equipment/grounding, group_b: drawings/cable_routes/fire_safety/automation, group_c: norms, drawings) → синтез → оптимизация"

    # Raw from grouped agents
    group_files = ["partial_group_a.json", "partial_group_b.json", "partial_group_c.json", "partial_drawings.json"]
    raw_total = 0
    agents = {}
    for gf in group_files:
        path = OUTPUT / gf
        data = load_json(path)
        if data is None:
            # попробуем сырой подсчёт по temp_id (если JSON битый)
            try:
                raw_text = path.read_text(encoding="utf-8")
                cnt = raw_text.count('"temp_id"')
                name = gf.replace("partial_", "").replace(".json", "")
                agents[name] = cnt
                raw_total += cnt
            except Exception:
                pass
            continue
        name = data.get("agent", gf.replace("partial_", "").replace(".json", ""))
        covers = data.get("covers", [])
        findings = data.get("findings", [])
        agents[name] = len(findings)
        raw_total += len(findings)
        if covers:
            agents[name] = f"{len(findings)} ({', '.join(covers)})"

    run["raw_from_agents"] = raw_total
    run["r1_agents"] = len(agents)
    run["agent_names"] = ", ".join(sorted(agents.keys()))
    run["agents_detail"] = agents

    # Findings — берём из audit_report_04.04.2026.xlsx (нет findings.json для этого прогона)
    # Парсим Excel для подсчёта
    findings_list = []
    try:
        from openpyxl import load_workbook
        wb_rpt = load_workbook(OUTPUT / "audit_report_04.04.2026.xlsx")
        ws_rpt = wb_rpt["133_23-ГК-ГРЩ"]
        cats = {}
        for row in range(3, ws_rpt.max_row + 1):
            cat_val = ws_rpt.cell(row=row, column=6).value or ""
            title_val = ws_rpt.cell(row=row, column=3).value or ""
            sheet_val = ws_rpt.cell(row=row, column=2).value or ""
            num = ws_rpt.cell(row=row, column=1).value
            if num is None or not str(num).isdigit():
                continue
            # Нормализуем категорию
            cat_clean = cat_val.replace("🔴 ", "").replace("🟠 ", "").replace("🟡 ", "").strip()
            cats[cat_clean] = cats.get(cat_clean, 0) + 1
            findings_list.append({
                "id": f"F-{int(num):03d}",
                "category": cat_clean,
                "title": title_val,
                "page": "",
                "sheet": sheet_val,
                "confidence": "",
                "source_agents": [],
            })
    except Exception as e:
        print(f"  Ошибка чтения Excel v2: {e}")

    run["findings_count"] = len(findings_list)
    run["findings_by_cat"] = {}
    for f in findings_list:
        c = f["category"]
        run["findings_by_cat"][c] = run["findings_by_cat"].get(c, 0) + 1
    run["avg_confidence"] = 0  # нет данных в Excel
    run["findings_list"] = findings_list

    # Optimization — из того же Excel
    run["opt_count"] = 0
    run["opt_accepted"] = "н/д"
    run["opt_accepted_pct"] = "н/д"
    run["opt_verdicts"] = {}
    try:
        if "ОПТ 133_23-ГК-ГРЩ" in wb_rpt.sheetnames:
            ws_opt = wb_rpt["ОПТ 133_23-ГК-ГРЩ"]
            opt_count = 0
            for row in range(3, ws_opt.max_row + 1):
                if ws_opt.cell(row=row, column=1).value and str(ws_opt.cell(row=row, column=1).value).isdigit():
                    opt_count += 1
            run["opt_count"] = opt_count
    except Exception:
        pass

    run["critics_count"] = 0
    run["critics_names"] = "нет (прямой синтез)"
    run["rejected_count"] = 0

    # Время
    run["wall_clock_min"] = "~9ч (04.04 09:01 → 18:50)"
    run["models"] = "Claude Sonnet 4.6 (R1 группы), Claude Opus 4.6 (синтез)"
    run["total_input_tokens"] = "н/д (Claude Code)"
    run["total_output_tokens"] = "н/д (Claude Code)"
    run["stages_info"] = {}

    return run


def collect_run4():
    """Run 4: 06.04.2026 — Claude Code, 12 индивидуальных агентов (данные в _output/)."""
    run = {"name": "Run 4 (06.04.2026)", "date": "06.04.2026"}
    run["architecture"] = "Claude Code: 12 R1-агентов → 1 критик → синтез → оптимизация + opt_critic"

    # Findings
    f_data = load_json(OUTPUT / "findings.json")
    if f_data:
        findings = f_data.get("findings", [])
        run["findings_count"] = len(findings)
        run["findings_by_cat"] = count_by_category(findings)
        run["avg_confidence"] = avg_confidence(findings)
        run["findings_list"] = findings
        stats = f_data.get("stats", {})
        run["stats"] = stats
    else:
        run["findings_count"] = 0
        run["findings_by_cat"] = {}
        run["avg_confidence"] = 0
        run["findings_list"] = []

    # Raw from agents (partial_*.json, исключая partial_group_*)
    raw_total = 0
    agents = {}
    for p in sorted(OUTPUT.glob("partial_*.json")):
        name = p.stem.replace("partial_", "")
        if name.startswith("group_"):
            continue  # пропускаем промежуточные группы
        data = load_json(p)
        if data is None:
            continue
        if isinstance(data, dict):
            fl = data.get("findings", [])
        elif isinstance(data, list):
            fl = data
        else:
            fl = []
        agents[name] = len(fl)
        raw_total += len(fl)
    run["raw_from_agents"] = raw_total
    run["r1_agents"] = len(agents)
    run["agent_names"] = ", ".join(sorted(agents.keys()))
    run["agents_detail"] = agents

    # Rejected
    rej = load_json(OUTPUT / "rejected.json")
    if rej:
        if isinstance(rej, dict):
            run["rejected_count"] = rej.get("total", len(rej.get("rejected", [])))
        elif isinstance(rej, list):
            run["rejected_count"] = len(rej)
        else:
            run["rejected_count"] = 0
    else:
        run["rejected_count"] = 0

    # Critics
    run["critics_count"] = 1
    run["critics_names"] = "review.json (Claude Code)"

    # Optimization
    opt = load_json(OUTPUT / "optimization.json")
    if opt:
        opts = opt.get("optimizations", opt.get("proposals", []))
        if isinstance(opts, list):
            run["opt_count"] = len(opts)
        else:
            run["opt_count"] = 0
    else:
        run["opt_count"] = 0

    opt_rev = load_json(OUTPUT / "optimization_review.json")
    if opt_rev:
        verdicts = count_opt_verdicts(opt_rev)
        accepted = verdicts.get("pass", 0) + verdicts.get("pass_with_conditions", 0)
        run["opt_accepted"] = accepted
        run["opt_verdicts"] = verdicts
        run["opt_accepted_pct"] = f"{round(accepted / run['opt_count'] * 100)}%" if run["opt_count"] else "—"
    else:
        run["opt_accepted"] = "н/д"
        run["opt_accepted_pct"] = "н/д"
        run["opt_verdicts"] = {}

    # Время
    run["wall_clock_min"] = "~5ч 21м (06.04 12:25 → 17:46)"
    run["models"] = "Claude Opus 4.6 (R1 + оптимизация), Claude Sonnet 4.6 (критик, нормы)"
    run["total_input_tokens"] = "н/д (Claude Code)"
    run["total_output_tokens"] = "н/д (Claude Code)"
    run["stages_info"] = {}

    return run


REFERENCE_XLSX = SCRIPT_DIR / "Сравнение_двух_вариатов_работы_ИИ__4__updated.xlsx"


def collect_run_reference():
    """Эталонный анализ — ручная проверка эксперта."""
    run = {"name": "Эталон (ручная проверка)", "date": "2026-04"}
    run["architecture"] = "Ручная экспертная проверка всех 47 замечаний из прогонов 1-2"

    findings_all = []
    findings_accepted = []

    try:
        from openpyxl import load_workbook
        wb_ref = load_workbook(REFERENCE_XLSX)
        ws = wb_ref["Сводные вариант ошибок"]
        for row in range(2, ws.max_row + 1):
            num = ws.cell(row=row, column=1).value
            if not str(num).isdigit():
                continue
            sheet_val = ws.cell(row=row, column=2).value or ""
            problem = ws.cell(row=row, column=3).value or ""
            severity = ws.cell(row=row, column=6).value or ""
            manual = str(ws.cell(row=row, column=11).value or "").strip()
            is_accepted = manual.lower().startswith("оставляем")

            entry = {
                "id": f"REF-{int(num):03d}",
                "category": severity,
                "title": problem,
                "page": "",
                "sheet": sheet_val,
                "confidence": "",
                "source_agents": [],
                "verdict": "accepted" if is_accepted else "rejected",
                "expert_comment": manual[:200],
            }
            findings_all.append(entry)
            if is_accepted:
                findings_accepted.append(entry)
    except Exception as e:
        print(f"  Ошибка чтения эталона: {e}")

    run["findings_count"] = len(findings_accepted)
    run["findings_list"] = findings_accepted
    run["raw_from_agents"] = len(findings_all)
    run["rejected_count"] = len(findings_all) - len(findings_accepted)
    run["all_findings"] = findings_all  # для расчёта precision/recall

    cats = {}
    for f in findings_accepted:
        c = f["category"]
        cats[c] = cats.get(c, 0) + 1
    run["findings_by_cat"] = cats
    run["avg_confidence"] = "—"

    run["r1_agents"] = "—"
    run["agent_names"] = "эксперт (ручная проверка)"
    run["critics_count"] = "—"
    run["critics_names"] = "—"
    run["opt_count"] = 0
    run["opt_accepted"] = "—"
    run["opt_accepted_pct"] = "—"
    run["opt_verdicts"] = {}
    run["wall_clock_min"] = "—"
    run["models"] = "человек"
    run["total_input_tokens"] = "—"
    run["total_output_tokens"] = "—"
    run["stages_info"] = {}

    return run


def compute_precision_recall(run, reference_run):
    """Вычисляет TP, FP, FN, Precision, Recall, F1 для прогона относительно эталона.

    Использует cross_match для определения: какие из 6 принятых эталонных замечаний
    данный прогон нашёл (TP), какие не нашёл (FN), сколько лишних (FP).
    """
    # Загружаем cross_match_v3
    cm_path = OUTPUT / "cross_match_v3.json"
    if not cm_path.exists():
        cm_path = OUTPUT / "cross_match_v2.json"
    cm = load_json(cm_path)
    if not cm:
        return {"tp": 0, "fp": 0, "fn": 0, "precision": 0, "recall": 0, "f1": 0}

    # Определяем ключ прогона в cross_match
    run_name = run.get("name", "")
    if "Run 1" in run_name:
        rk = "run1"
    elif "Run 2" in run_name:
        rk = "run2"
    elif "v2" in run_name or "Run 3 v2" in run_name:
        rk = "run3v2"
    elif "Run 4" in run_name:
        rk = "run3"  # run3 в cross_match = Run 4
    else:
        return {"tp": 0, "fp": 0, "fn": 0, "precision": 0, "recall": 0, "f1": 0}

    tp = 0
    fp = 0
    fn = 0

    for item in cm:
        has_run = item.get(rk) is not None
        ref = item.get("reference")
        is_accepted = ref and ref.get("verdict") == "accepted"

        if has_run and is_accepted:
            tp += 1
        elif has_run and not is_accepted:
            fp += 1
        elif not has_run and is_accepted:
            fn += 1

    precision = round(tp / (tp + fp), 3) if (tp + fp) > 0 else 0
    recall = round(tp / (tp + fn), 3) if (tp + fn) > 0 else 0
    f1 = round(2 * precision * recall / (precision + recall), 3) if (precision + recall) > 0 else 0

    return {"tp": tp, "fp": fp, "fn": fn, "precision": precision, "recall": recall, "f1": f1}


# ─── Генерация Excel ────────────────────────────────────────────────────────

def generate_excel(runs):
    wb = Workbook()

    # ═══════════════════════════════════════════════════════════════════════
    # ЛИСТ 1: СВОДКА ПРОГОНОВ
    # ═══════════════════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Сводка прогонов"
    ws1.sheet_properties.tabColor = "2F5496"

    # Заголовки: Метрика | Run 1 | Run 2 | Run 3
    headers = ["Метрика"] + [r["name"] for r in runs]
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        style_cell(cell, HEADER_FILL, HEADER_FONT, "center")

    # Строки метрик
    metrics = [
        ("Дата прогона", [r["date"] for r in runs]),
        ("Архитектура", [r["architecture"] for r in runs]),
        ("", [None] * len(runs)),  # разделитель
        ("АГЕНТЫ И КРИТИКИ", [None] * len(runs)),
        ("Кол-во R1-агентов", [r["r1_agents"] for r in runs]),
        ("Имена агентов", [r["agent_names"] for r in runs]),
        ("Кол-во критиков", [r["critics_count"] for r in runs]),
        ("Критики", [r["critics_names"] for r in runs]),
        ("", [None] * len(runs)),
        ("ВРЕМЯ", [None] * len(runs)),
        ("Wall-clock", [r.get("wall_clock_min", "н/д") for r in runs]),
        ("Модели", [r.get("models", "н/д") for r in runs]),
        ("", [None] * len(runs)),
        ("ТОКЕНЫ", [None] * len(runs)),
        ("Input tokens", [r.get("total_input_tokens", "н/д") for r in runs]),
        ("Output tokens", [r.get("total_output_tokens", "н/д") for r in runs]),
    ]

    # Детали по этапам для Run 1
    if runs[0].get("stages_info"):
        metrics.append(("", [None] * len(runs)))
        metrics.append(("ЭТАПЫ (Run 1 — pipeline_log)", [None] * len(runs)))
        for stage_name, info in runs[0]["stages_info"].items():
            dur = info["duration_sec"]
            dur_str = f"{dur/60:.1f} мин" if dur >= 60 else f"{dur:.0f} сек"
            tok_str = f"in:{info['input_tokens']:,} / out:{info['output_tokens']:,}" if info["input_tokens"] else "—"
            val = f"{dur_str} | {info['model']} | {tok_str}"
            metrics.append((f"  {stage_name}", [val, "—", "—"]))

    row = 2
    for label, values in metrics:
        if not label and all(v is None for v in values):
            row += 1
            continue
        if label and all(v is None for v in values):
            # Секция
            cell = ws1.cell(row=row, column=1, value=label)
            style_cell(cell, SECTION_FILL, SECTION_FONT)
            for col in range(2, len(headers) + 1):
                style_cell(ws1.cell(row=row, column=col), SECTION_FILL)
            row += 1
            continue
        cell = ws1.cell(row=row, column=1, value=label)
        style_cell(cell, LABEL_FILL, LABEL_FONT, wrap=True)
        for col, v in enumerate(values, 2):
            c = ws1.cell(row=row, column=col, value=v if v is not None else "")
            style_cell(c, wrap=True)
        row += 1

    ws1.column_dimensions["A"].width = 30
    for ci in range(2, len(headers) + 1):
        ws1.column_dimensions[get_column_letter(ci)].width = 40
    ws1.freeze_panes = "B2"

    # ═══════════════════════════════════════════════════════════════════════
    # ЛИСТ 2: КАЧЕСТВО
    # ═══════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Качество")
    ws2.sheet_properties.tabColor = "548235"

    for col, h in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        style_cell(cell, HEADER_FILL, HEADER_FONT, "center")

    quality_metrics = [
        ("ЗАМЕЧАНИЯ", [None] * len(runs)),
        ("Замечаний от агентов (raw)", [r["raw_from_agents"] for r in runs]),
        ("Отклонено критиком", [r["rejected_count"] for r in runs]),
        ("Финальных замечаний", [r["findings_count"] for r in runs]),
        ("", [None] * len(runs)),
        ("ПО КАТЕГОРИЯМ", [None] * len(runs)),
        ("Критическое", [r["findings_by_cat"].get("Критическое", r["findings_by_cat"].get("КРИТИЧЕСКОЕ", 0)) for r in runs]),
        ("Экономическое", [r["findings_by_cat"].get("Экономическое", r["findings_by_cat"].get("ЭКОНОМИЧЕСКОЕ", 0)) for r in runs]),
        ("Эксплуатационное", [r["findings_by_cat"].get("Эксплуатационное", 0) for r in runs]),
        ("Рекомендательное", [r["findings_by_cat"].get("Рекомендательное", r["findings_by_cat"].get("РЕКОМЕНДАТЕЛЬНОЕ", 0)) for r in runs]),
        ("Проверить по смежным", [r["findings_by_cat"].get("ПРОВЕРИТЬ ПО СМЕЖНЫМ", 0) for r in runs]),
        ("", [None] * len(runs)),
        ("CONFIDENCE", [None] * len(runs)),
        ("Средний confidence", [r["avg_confidence"] for r in runs]),
        ("", [None] * len(runs)),
        ("ОПТИМИЗАЦИЯ", [None] * len(runs)),
        ("Предложений оптимизации", [r["opt_count"] for r in runs]),
        ("Принято (pass + pass_with_conditions)", [r["opt_accepted"] for r in runs]),
        ("% принятых", [r["opt_accepted_pct"] for r in runs]),
    ]

    # Детали вердиктов оптимизации для Run 2 и Run 3
    all_verdicts = set()
    for r in runs:
        all_verdicts.update(r.get("opt_verdicts", {}).keys())
    if all_verdicts:
        quality_metrics.append(("", [None] * len(runs)))
        quality_metrics.append(("ВЕРДИКТЫ ОПТИМИЗАЦИИ", [None] * len(runs)))
        for v in sorted(all_verdicts):
            quality_metrics.append((f"  {v}", [r.get("opt_verdicts", {}).get(v, 0) for r in runs]))

    # Детали по агентам для Run 3 v2
    for ri, r in enumerate(runs):
        if r.get("agents_detail"):
            quality_metrics.append(("", [None] * len(runs)))
            quality_metrics.append((f"ЗАМЕЧАНИЙ ПО АГЕНТАМ ({r['name']})", [None] * len(runs)))
            for agent, count in sorted(r["agents_detail"].items()):
                row_vals = ["—"] * len(runs)
                row_vals[ri] = count
                quality_metrics.append((f"  {agent}", row_vals))

    # Precision / Recall / F1 относительно эталона
    ref_run = next((r for r in runs if "Эталон" in r.get("name", "")), None)
    if ref_run:
        quality_metrics.append(("", [None] * len(runs)))
        quality_metrics.append(("ТОЧНОСТЬ (vs эталон)", [None] * len(runs)))
        pr_data = [compute_precision_recall(r, ref_run) if "Эталон" not in r.get("name", "") else {} for r in runs]
        quality_metrics.append(("TP (верно найдены)", [d.get("tp", "—") for d in pr_data]))
        quality_metrics.append(("FP (ложные)", [d.get("fp", "—") for d in pr_data]))
        quality_metrics.append(("FN (пропущены)", [d.get("fn", "—") for d in pr_data]))
        quality_metrics.append(("Precision (TP/(TP+FP))", [d.get("precision", "—") for d in pr_data]))
        quality_metrics.append(("Recall (TP/(TP+FN))", [d.get("recall", "—") for d in pr_data]))
        quality_metrics.append(("F1-score", [d.get("f1", "—") for d in pr_data]))

    row = 2
    for label, values in quality_metrics:
        if not label and all(v is None for v in values):
            row += 1
            continue
        if label and all(v is None for v in values):
            cell = ws2.cell(row=row, column=1, value=label)
            style_cell(cell, SECTION_FILL, SECTION_FONT)
            for col in range(2, len(headers) + 1):
                style_cell(ws2.cell(row=row, column=col), SECTION_FILL)
            row += 1
            continue

        cell = ws2.cell(row=row, column=1, value=label)
        is_cat = label in CAT_FILLS
        style_cell(cell, CAT_FILLS.get(label, LABEL_FILL), LABEL_FONT if not is_cat else Font(bold=True, size=10), wrap=True)
        for col, v in enumerate(values, 2):
            c = ws2.cell(row=row, column=col, value=v if v is not None else "")
            fill = CAT_FILLS.get(label) if is_cat else None
            style_cell(c, fill=fill, align_h="center")
        row += 1

    ws2.column_dimensions["A"].width = 36
    for ci in range(2, len(headers) + 1):
        ws2.column_dimensions[get_column_letter(ci)].width = 22
    ws2.freeze_panes = "B2"

    # ═══════════════════════════════════════════════════════════════════════
    # ЛИСТ 3: ДЕТАЛИ ЗАМЕЧАНИЙ (горизонтально, прогоны рядом)
    # ═══════════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Детали замечаний")
    ws3.sheet_properties.tabColor = "BF8F00"

    # Столбцы одного блока: ID, Категория, Заголовок, Стр., Лист, Confidence, Агент
    BLOCK_HEADERS = ["ID", "Категория", "Заголовок", "Стр.", "Лист", "Confidence", "Агент"]
    BLOCK_WIDTHS = [8, 16, 35, 6, 6, 11, 18]
    SEP_WIDTH = 2  # ширина пустого столбца-разделителя

    col_offset = 1
    for run_idx, run in enumerate(runs):
        # Заголовок прогона (merged)
        merge_start = col_offset
        merge_end = col_offset + len(BLOCK_HEADERS) - 1
        title_cell = ws3.cell(row=1, column=merge_start, value=run["name"])
        style_cell(title_cell, RUN_HEADER_FILL, RUN_HEADER_FONT, "center")
        for c in range(merge_start + 1, merge_end + 1):
            style_cell(ws3.cell(row=1, column=c), RUN_HEADER_FILL)
        ws3.merge_cells(start_row=1, start_column=merge_start, end_row=1, end_column=merge_end)

        # Заголовки столбцов
        for hi, h in enumerate(BLOCK_HEADERS):
            cell = ws3.cell(row=2, column=col_offset + hi, value=h)
            style_cell(cell, HEADER_FILL, HEADER_FONT, "center")

        # Данные
        findings = run.get("findings_list", [])
        for fi, f in enumerate(findings):
            r = fi + 3  # строка данных начинается с 3
            fid = f.get("id", "")
            cat = f.get("category") or f.get("severity", "")
            title = f.get("title") or f.get("problem", "")
            page = f.get("page", "")
            if isinstance(page, list):
                page = ", ".join(str(p) for p in page)
            sheet = f.get("sheet", "")
            if isinstance(sheet, list):
                sheet = ", ".join(str(s) for s in sheet)
            conf = f.get("confidence") or f.get("practicality_score", "")
            if isinstance(conf, float):
                conf = round(conf, 2)
            agent = ", ".join(f.get("source_agents", [])) if f.get("source_agents") else ""

            vals = [fid, cat, title, page, sheet, conf, agent]
            for vi, v in enumerate(vals):
                c = ws3.cell(row=r, column=col_offset + vi, value=v)
                c.border = THIN_BORDER
                c.alignment = Alignment(vertical="center", wrap_text=(vi == 2))
            # Подсветка категории
            cat_cell = ws3.cell(row=r, column=col_offset + 1)
            if cat in CAT_FILLS:
                cat_cell.fill = CAT_FILLS[cat]

        # Ширины столбцов блока
        for wi, w in enumerate(BLOCK_WIDTHS):
            ws3.column_dimensions[get_column_letter(col_offset + wi)].width = w

        # Пустой столбец-разделитель (если не последний прогон)
        col_offset += len(BLOCK_HEADERS)
        if run_idx < len(runs) - 1:
            ws3.column_dimensions[get_column_letter(col_offset)].width = SEP_WIDTH
            col_offset += 1

    ws3.freeze_panes = "A3"

    # ═══════════════════════════════════════════════════════════════════════
    # ЛИСТ 4: СОПОСТАВЛЕНИЕ ЗАМЕЧАНИЙ (cross-match, 4 прогона)
    # ═══════════════════════════════════════════════════════════════════════
    ws_cm = wb.create_sheet("Сопоставление замечаний")
    ws_cm.sheet_properties.tabColor = "C55A11"

    for cm_name in ("cross_match_v3.json", "cross_match_v2.json", "cross_match.json"):
        cross_match_path = OUTPUT / cm_name
        if cross_match_path.exists():
            break
    cross_match = load_json(cross_match_path)

    RUN_KEYS = ["run1", "run2", "run3v2", "run3", "reference"]  # порядок: Run1, Run2, Run3v2, Run4(=run3), Эталон
    RUN_LABELS = ["Run 1", "Run 2", "Run 3 v2", "Run 4", "Эталон"]
    NUM_RUNS = len(RUN_KEYS)

    MATCH_FILLS = {
        5: PatternFill(start_color="00B050", end_color="00B050", fill_type="solid"),  # тёмно-зелёный — все 5
        4: PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),  # зелёный — 4 из 5
        3: PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid"),  # голубой — 3 из 5
        2: PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),  # жёлтый — 2 из 5
        1: PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),  # красный — 1 из 5
    }
    ROW_ACCEPTED_FILL = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")  # синий
    ROW_REJECTED_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")  # серый
    MATCH_FONTS = {
        4: Font(bold=True, color="FFFFFF"),
    }

    if cross_match:
        # Заголовки: №, Тема, [sep], Run1(ID,Cat,Desc), [sep], Run2(...), [sep], Run3v2(...), [sep], Run4(...), Совпадений
        # Загрузить экспертные комментарии из эталонного Excel
        expert_comments = {}  # ref_num → comment_text
        try:
            from openpyxl import load_workbook as _lw
            wb_ref = _lw(REFERENCE_XLSX)
            ws_ref = wb_ref["Сводные вариант ошибок"]
            for erow in range(2, ws_ref.max_row + 1):
                enum = ws_ref.cell(row=erow, column=1).value
                if str(enum).isdigit():
                    manual = str(ws_ref.cell(row=erow, column=11).value or "").strip()
                    expert_comments[int(enum)] = manual
        except Exception:
            pass

        cm_headers = ["№", "Тема замечания", ""]
        run_col_map = {}  # run_key → (id_col, cat_col, desc_col)
        col_idx = 4
        for ri, (rk, rl) in enumerate(zip(RUN_KEYS, RUN_LABELS)):
            run_col_map[rk] = (col_idx, col_idx + 1, col_idx + 2)
            cm_headers += [f"{rl}\nID", f"{rl}\nКатегория", f"{rl}\nОписание"]
            col_idx += 3
            if ri < NUM_RUNS - 1:
                cm_headers.append("")  # разделитель
                col_idx += 1
        cm_headers.append(f"Совпадений\n(из {NUM_RUNS})")
        cm_headers.append("Экспертная оценка")
        total_cols = len(cm_headers)
        match_col = total_cols - 1  # предпоследний = совпадений
        expert_col = total_cols  # последний = экспертная оценка

        for col, h in enumerate(cm_headers, 1):
            if h:
                cell = ws_cm.cell(row=1, column=col, value=h)
                style_cell(cell, HEADER_FILL, HEADER_FONT, "center", wrap=True)

        def match_count(item):
            return sum(1 for k in RUN_KEYS if item.get(k))

        sorted_matches = sorted(cross_match, key=lambda x: (-match_count(x), x.get("topic", "")))

        for i, item in enumerate(sorted_matches, 1):
            row = i + 1
            mc = match_count(item)

            ws_cm.cell(row=row, column=1, value=i)
            ws_cm.cell(row=row, column=2, value=item.get("topic", ""))

            for rk, (id_c, cat_c, desc_c) in run_col_map.items():
                run_data = item.get(rk)
                if run_data:
                    ws_cm.cell(row=row, column=id_c, value=run_data.get("id", ""))
                    cat_val = run_data.get("category", "")
                    cat_cell = ws_cm.cell(row=row, column=cat_c, value=cat_val)
                    if cat_val in CAT_FILLS:
                        cat_cell.fill = CAT_FILLS[cat_val]
                    ws_cm.cell(row=row, column=desc_c, value=run_data.get("title_short", ""))
                else:
                    ws_cm.cell(row=row, column=id_c, value="—")
                    ws_cm.cell(row=row, column=cat_c, value="—")
                    ws_cm.cell(row=row, column=desc_c, value="—")

            mc_cell = ws_cm.cell(row=row, column=match_col, value=mc)
            if mc in MATCH_FILLS:
                mc_cell.fill = MATCH_FILLS[mc]
            if mc in MATCH_FONTS:
                mc_cell.font = MATCH_FONTS[mc]

            # Столбец X: экспертная оценка
            ref = item.get("reference")
            expert_text = ""
            row_fill = None
            if ref:
                ref_id = ref.get("id", "")
                verdict = ref.get("verdict", "")
                try:
                    ref_num = int(ref_id.replace("REF-", ""))
                    expert_text = expert_comments.get(ref_num, "")
                except (ValueError, AttributeError):
                    pass
                if verdict == "accepted":
                    row_fill = ROW_ACCEPTED_FILL
                else:
                    row_fill = ROW_REJECTED_FILL

            expert_cell = ws_cm.cell(row=row, column=expert_col, value=expert_text)
            expert_cell.border = THIN_BORDER
            expert_cell.alignment = Alignment(vertical="center", wrap_text=True)

            # Borders, alignment, and row coloring
            for col in range(1, total_cols + 1):
                c = ws_cm.cell(row=row, column=col)
                h = cm_headers[col - 1] if col <= len(cm_headers) else ""
                if h:  # не пустой разделитель
                    c.border = THIN_BORDER
                    # Подсветка всей строки: синий (принято) / серый (отклонено)
                    if row_fill and col != match_col:
                        # Не перезаписываем fill у ячеек категорий (они уже раскрашены)
                        existing_cat = c.value
                        if existing_cat not in CAT_FILLS and col not in (match_col,):
                            c.fill = row_fill
                c.alignment = Alignment(vertical="center", wrap_text=(h and ("Описание" in h or col == 2 or "оценка" in h)))

        # Сводка
        total_row = len(sorted_matches) + 3
        ws_cm.cell(row=total_row, column=2, value=f"Уникальных тем: {len(sorted_matches)}")
        style_cell(ws_cm.cell(row=total_row, column=2), SECTION_FILL, SECTION_FONT)

        for mc_val in range(NUM_RUNS, 0, -1):
            cnt = sum(1 for x in sorted_matches if match_count(x) == mc_val)
            label = f"Найдено {mc_val} из {NUM_RUNS}: {cnt}"
            r = total_row + (NUM_RUNS - mc_val + 1)
            ws_cm.cell(row=r, column=2, value=label)
            style_cell(ws_cm.cell(row=r, column=2), MATCH_FILLS.get(mc_val))
            if mc_val in MATCH_FONTS:
                ws_cm.cell(row=r, column=2).font = MATCH_FONTS[mc_val]

        # Уникальных по прогонам
        r_uniq = total_row + NUM_RUNS + 2
        for rk, rl in zip(RUN_KEYS, RUN_LABELS):
            cnt = sum(1 for x in sorted_matches if match_count(x) == 1 and x.get(rk))
            ws_cm.cell(row=r_uniq, column=2, value=f"Уникальных для {rl}: {cnt}")
            r_uniq += 1

    # Ширины столбцов
    cm_widths = [4, 50, 2]  # №, тема, sep
    for ri in range(NUM_RUNS):
        cm_widths += [8, 16, 40]  # ID, cat, desc
        if ri < NUM_RUNS - 1:
            cm_widths.append(2)  # sep
    cm_widths.append(12)  # совпадений
    cm_widths.append(50)  # экспертная оценка
    for i, w in enumerate(cm_widths, 1):
        ws_cm.column_dimensions[get_column_letter(i)].width = w
    ws_cm.freeze_panes = "C2"
    if cross_match:
        ws_cm.auto_filter.ref = f"A1:{get_column_letter(expert_col)}{len(cross_match) + 1}"

    # ═══════════════════════════════════════════════════════════════════════
    # ЛИСТ 4: ХРОНОЛОГИЯ (шаблон для будущих прогонов)
    # ═══════════════════════════════════════════════════════════════════════
    ws4 = wb.create_sheet("Хронология")
    ws4.sheet_properties.tabColor = "7030A0"

    chrono_headers = [
        "№ прогона", "Дата", "Архитектура / изменения",
        "R1-агентов", "Критиков",
        "Raw замечаний", "Финальных замечаний",
        "Критич.", "Эконом.", "Эксплуат.",
        "Ср. confidence",
        "Оптимизаций", "Принято (%)",
        "Wall-clock",
        "Комментарий",
    ]
    for col, h in enumerate(chrono_headers, 1):
        cell = ws4.cell(row=1, column=col, value=h)
        style_cell(cell, HEADER_FILL, HEADER_FONT, "center", wrap=True)

    for i, run in enumerate(runs, 1):
        row = i + 1
        cats = run["findings_by_cat"]
        ws4.cell(row=row, column=1, value=i)
        ws4.cell(row=row, column=2, value=run["date"])
        ws4.cell(row=row, column=3, value=run["architecture"])
        ws4.cell(row=row, column=4, value=run["r1_agents"])
        ws4.cell(row=row, column=5, value=run["critics_count"])
        ws4.cell(row=row, column=6, value=run["raw_from_agents"])
        ws4.cell(row=row, column=7, value=run["findings_count"])
        ws4.cell(row=row, column=8, value=cats.get("Критическое", cats.get("КРИТИЧЕСКОЕ", 0)))
        ws4.cell(row=row, column=9, value=cats.get("Экономическое", cats.get("ЭКОНОМИЧЕСКОЕ", 0)))
        ws4.cell(row=row, column=10, value=cats.get("Эксплуатационное", 0))
        ws4.cell(row=row, column=11, value=run["avg_confidence"])
        ws4.cell(row=row, column=12, value=run["opt_count"])
        ws4.cell(row=row, column=13, value=run["opt_accepted_pct"])
        ws4.cell(row=row, column=14, value=run.get("wall_clock_min", "н/д"))
        ws4.cell(row=row, column=15, value="")  # комментарий

        for col in range(1, len(chrono_headers) + 1):
            c = ws4.cell(row=row, column=col)
            c.border = THIN_BORDER
            c.alignment = Alignment(vertical="center", wrap_text=col in (3, 15))

    # Пустые строки для будущих прогонов (ещё 10)
    for j in range(len(runs) + 2, len(runs) + 12):
        ws4.cell(row=j, column=1, value=j - 1)
        for col in range(1, len(chrono_headers) + 1):
            ws4.cell(row=j, column=col).border = THIN_BORDER

    col_widths4 = [10, 16, 50, 12, 10, 14, 14, 10, 10, 12, 14, 13, 13, 18, 40]
    for i, w in enumerate(col_widths4, 1):
        ws4.column_dimensions[get_column_letter(i)].width = w
    ws4.freeze_panes = "A2"

    # ═══ Сохранение ═══
    wb.save(str(EXCEL_PATH))
    print(f"Excel сохранён: {EXCEL_PATH}")
    return EXCEL_PATH


# ─── Main ────────────────────────────────────────────────────────────────────

def main():
    print(f"Проект: {PROJECT_DIR}")
    print(f"_output:     {OUTPUT}")
    print(f"_output_old: {OUTPUT_OLD}")
    print()

    if not PROJECT_DIR.exists():
        print(f"ОШИБКА: папка проекта не найдена: {PROJECT_DIR}")
        sys.exit(1)

    runs = []

    print("Сбор данных Run 1 (27.03.2026 — Python-пайплайн)...")
    runs.append(collect_run1())
    print(f"  Замечаний: {runs[-1]['findings_count']}, confidence: {runs[-1]['avg_confidence']}")

    print("Сбор данных Run 2 (28-29.03.2026 — 6 агентов)...")
    runs.append(collect_run2())
    print(f"  Замечаний: {runs[-1]['findings_count']}, confidence: {runs[-1]['avg_confidence']}")

    print("Сбор данных Run 3 v2 (04.04.2026 — группированные агенты, Sonnet+Opus)...")
    runs.append(collect_run3_v2())
    print(f"  Замечаний: {runs[-1]['findings_count']}, confidence: {runs[-1]['avg_confidence']}")

    print("Сбор данных Run 4 (06.04.2026 — 12 индивидуальных агентов)...")
    runs.append(collect_run4())
    print(f"  Замечаний: {runs[-1]['findings_count']}, confidence: {runs[-1]['avg_confidence']}")

    print("Сбор эталонных данных (ручная проверка эксперта)...")
    runs.append(collect_run_reference())
    print(f"  Принято: {runs[-1]['findings_count']}, отклонено: {runs[-1]['rejected_count']}")

    print()
    generate_excel(runs)


if __name__ == "__main__":
    main()
