import sys, io, json
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

with open('projects/EOM/133_23-ГК-ГРЩ/_output/findings.json', 'r', encoding='utf-8') as f:
    fdata = json.load(f)
with open('projects/EOM/133_23-ГК-ГРЩ/_output/optimization.json', 'r', encoding='utf-8') as f:
    odata = json.load(f)
with open('projects/EOM/133_23-ГК-ГРЩ/_output/optimization_review.json', 'r', encoding='utf-8') as f:
    ordata = json.load(f)

findings = fdata['findings']
proposals = odata.get('proposals', [])
opt_verdicts = {r['opt_id']: r for r in ordata.get('reviews', [])}

wb = Workbook()
now = datetime.now().strftime('%d.%m.%Y %H:%M')
project_id = '133_23-ГК-ГРЩ'
thin = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))

HEADER_FILL = PatternFill('solid', fgColor='1F497D')
HEADER_FONT = Font(bold=True, color='FFFFFF', size=10)
TITLE_FILL = PatternFill('solid', fgColor='2E75B6')
TITLE_FONT = Font(bold=True, color='FFFFFF', size=11)
TOTAL_FILL = PatternFill('solid', fgColor='2E3F50')

CAT_MAP = {
    'Критическое':      {'emoji': '\U0001f534', 'label': 'КРИТИЧЕСКОЕ',      'row_fill': 'FFCCCC', 'cell_fill': 'FFD9D9', 'font_color': 'C00000'},
    'Экономическое':    {'emoji': '\U0001f7e0', 'label': 'ЭКОНОМИЧЕСКОЕ',    'row_fill': 'FCE4D6', 'cell_fill': 'FDEBD9', 'font_color': 'C05000'},
    'Эксплуатационное': {'emoji': '\U0001f7e1', 'label': 'ЭКСПЛУАТАЦИОННОЕ', 'row_fill': 'FFFACD', 'cell_fill': 'FFFFB3', 'font_color': '806600'},
    'Рекомендательное': {'emoji': '\U0001f535', 'label': 'РЕКОМЕНДАТЕЛЬНОЕ', 'row_fill': 'DDEEFF', 'cell_fill': 'DDEEFF', 'font_color': '1F497D'},
}

def ci(cat):
    return CAT_MAP.get(cat, {'emoji': '\u26aa', 'label': cat.upper(), 'row_fill': 'F8F8F8', 'cell_fill': 'F8F8F8', 'font_color': '333333'})

counts = {}
for f in findings:
    counts[f['category']] = counts.get(f['category'], 0) + 1
cat_order = ['Критическое', 'Экономическое', 'Эксплуатационное', 'Рекомендательное']

# ===== СВОДКА =====
ws1 = wb.active
ws1.title = 'СВОДКА'
h1 = ['№', 'Проект (ID)', 'Объект / Раздел']
for cat in cat_order:
    h1.append(f'{ci(cat)["emoji"]} {cat[:6]}.')
h1.append('Итого')
w1 = [5, 30, 38, 14, 14, 16, 16, 10]
for i, (h, w) in enumerate(zip(h1, w1), 1):
    c = ws1.cell(row=1, column=i, value=h)
    c.font = HEADER_FONT; c.fill = HEADER_FILL; c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True); c.border = thin
    ws1.column_dimensions[get_column_letter(i)].width = w

ws1.merge_cells('A2:H2')
c = ws1.cell(row=2, column=1, value=f'СВОДНЫЙ ОТЧЁТ АУДИТА   |   Дата: {now}   |   Мультиагентный пайплайн')
c.font = TITLE_FONT; c.fill = TITLE_FILL
for col in range(1, 9): ws1.cell(row=2, column=col).border = thin

row3 = [1, project_id, 'ГРЩ / Электроснабжение (EM)'] + [counts.get(cat, 0) for cat in cat_order] + [len(findings)]
for i, v in enumerate(row3, 1):
    c = ws1.cell(row=3, column=i, value=v)
    c.font = Font(bold=(i in [2, 8]), size=10); c.alignment = Alignment(horizontal='center', vertical='center'); c.border = thin
    if 4 <= i <= 7: c.fill = PatternFill('solid', fgColor=ci(cat_order[i-4])['cell_fill'])

for col in range(1, 9):
    c = ws1.cell(row=4, column=col)
    c.fill = TOTAL_FILL; c.font = Font(bold=True, color='FFFFFF', size=10); c.alignment = Alignment(horizontal='center'); c.border = thin
ws1.cell(row=4, column=1, value='ИТОГО')
for i, cat in enumerate(cat_order, 4):
    ws1.cell(row=4, column=i, value=counts.get(cat, 0))
ws1.cell(row=4, column=8, value=len(findings))

# ===== ЗАМЕЧАНИЯ =====
ws2 = wb.create_sheet(project_id)
h2 = ['№', 'Лист/Раздел', 'Проблема', 'Описание', 'Решение', 'Категория', 'Чем грозит']
w2 = [5, 24, 28, 52, 48, 22, 32]
for i, (h, w) in enumerate(zip(h2, w2), 1):
    c = ws2.cell(row=1, column=i, value=h)
    c.font = HEADER_FONT; c.fill = HEADER_FILL; c.alignment = Alignment(horizontal='center' if i == 1 else 'left', vertical='center', wrap_text=True); c.border = thin
    ws2.column_dimensions[get_column_letter(i)].width = w

ws2.merge_cells('A2:G2')
c = ws2.cell(row=2, column=1, value=f'{project_id}  |  аудит от {now}  |  замечаний: {len(findings)}')
c.font = TITLE_FONT; c.fill = TITLE_FILL
for col in range(1, 8): ws2.cell(row=2, column=col).border = thin

cat_pri = {'Критическое': 0, 'Экономическое': 1, 'Эксплуатационное': 2, 'Рекомендательное': 3}
sf = sorted(findings, key=lambda f: cat_pri.get(f['category'], 9))

RISK = {
    'Критическое': 'Нарушение обязательных норм',
    'Экономическое': 'Финансовые потери при закупке/монтаже',
    'Эксплуатационное': 'Проблемы при эксплуатации объекта',
    'Рекомендательное': 'Несоответствие в документации',
}

for idx, f in enumerate(sf):
    r = idx + 3
    info = ci(f['category'])
    rf = PatternFill('solid', fgColor=info['row_fill'])
    sheet_ref = f'Лист {f["sheet"]}' if f.get('sheet') else f'Стр. {f.get("page", "")}'
    vals = [idx+1, sheet_ref, f.get('title',''), f.get('description',''), f.get('recommendation',''), f'{info["emoji"]} {info["label"]}', RISK.get(f['category'], '')]
    for col, v in enumerate(vals, 1):
        c = ws2.cell(row=r, column=col, value=v)
        c.fill = rf; c.border = thin
        c.alignment = Alignment(horizontal='center' if col == 1 else 'left', vertical='top', wrap_text=True)
        if col == 6: c.font = Font(bold=True, color=info['font_color'], size=10)
        elif col == 4: c.font = Font(size=9)
        elif col == 1: c.font = Font(bold=True, size=10)
        else: c.font = Font(size=10)

rf = len(sf) + 3
ws2.merge_cells(f'A{rf}:G{rf}')
ws2.cell(row=rf, column=1, value=f'Итого замечаний: {len(findings)}  |  Мультиагентный аудит  |  {now}').font = Font(italic=True, size=9, color='666666')
ws2.cell(row=rf, column=1).alignment = Alignment(horizontal='center')

# ===== ОПТИМИЗАЦИЯ =====
ws3 = wb.create_sheet(f'ОПТ {project_id}')
h3 = ['№', 'ID', 'Раздел/Лист', 'Текущее решение', 'Предложение', 'Тип', 'Вердикт', 'Привязка', 'Риски']
w3 = [5, 10, 24, 42, 42, 20, 16, 12, 32]
for i, (h, w) in enumerate(zip(h3, w3), 1):
    c = ws3.cell(row=1, column=i, value=h)
    c.font = HEADER_FONT; c.fill = HEADER_FILL; c.alignment = Alignment(horizontal='center' if i <= 2 else 'left', vertical='center', wrap_text=True); c.border = thin
    ws3.column_dimensions[get_column_letter(i)].width = w

passed = sum(1 for r in ordata.get('reviews', []) if 'pass' in r.get('verdict', ''))
ws3.merge_cells('A2:I2')
c = ws3.cell(row=2, column=1, value=f'{project_id}  |  оптимизаций: {len(proposals)} (принято: {passed})  |  отчёт: {now}')
c.font = TITLE_FONT; c.fill = TITLE_FILL
for col in range(1, 10): ws3.cell(row=2, column=col).border = thin

TYPE_E = {'замена_оборудования': '\U0001f4b0 Замена оборудования', 'оптимизация_кабеля': '\U0001f527 Оптимизация кабеля', 'упрощение_конструкции': '\U0001f527 Проще конструкция', 'исправление_по_findings': '\u26a0\ufe0f По замечаниям'}
VF = {'pass': PatternFill('solid', fgColor='D4EDDA'), 'pass_with_conditions': PatternFill('solid', fgColor='FFF3CD'), 'conflict_with_finding': PatternFill('solid', fgColor='F8D7DA'), 'reject_reliability': PatternFill('solid', fgColor='F8D7DA')}

for idx, p in enumerate(proposals):
    r = idx + 3
    oid = p.get('opt_id', '')
    rv = opt_verdicts.get(oid, {})
    verdict = rv.get('verdict', '')
    vals = [idx+1, oid, 'EM / ГРЩ', p.get('current',''), p.get('proposed',''), TYPE_E.get(p.get('type',''), p.get('type','')), verdict, p.get('finding_ref','') or '\u2014', p.get('risk','') or '\u2014']
    for col, v in enumerate(vals, 1):
        c = ws3.cell(row=r, column=col, value=v)
        c.font = Font(size=10, bold=(col <= 2)); c.alignment = Alignment(horizontal='center' if col <= 2 else 'left', vertical='top', wrap_text=True); c.border = thin
        if col == 7 and verdict in VF: c.fill = VF[verdict]

rf = len(proposals) + 3
ws3.merge_cells(f'A{rf}:I{rf}')
ws3.cell(row=rf, column=1, value=f'Итого: {len(proposals)} предложений  |  Opus  |  {now}').font = Font(italic=True, size=9, color='666666')
ws3.cell(row=rf, column=1).alignment = Alignment(horizontal='center')

out = 'projects/EOM/133_23-ГК-ГРЩ/_output/audit_report_28.03.2026.xlsx'
wb.save(out)
print(f'OK: {out}')
print(f'Findings: {len(findings)}, Optimizations: {len(proposals)}')
